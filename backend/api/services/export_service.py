"""
Export service for Google Sheets and Excel integration.

This module provides utilities for exporting analytics data to Google Sheets
and Excel files with proper formatting.
"""

from typing import List, Dict, Any, Optional
from datetime import date
import os
import logging
from io import BytesIO

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to various formats"""

    def __init__(self, google_credentials_path: Optional[str] = None):
        """
        Initialize export service.

        Args:
            google_credentials_path: Path to Google service account credentials JSON
        """
        self.google_credentials_path = google_credentials_path or os.getenv(
            'GOOGLE_CREDENTIALS_PATH'
        )
        self._sheets_service = None

    def _get_sheets_service(
        self, 
        access_token: Optional[str] = None, 
        refresh_token: Optional[str] = None,
        client_id: Optional[str] = None,
        client_secret: Optional[str] = None,
        token_uri: Optional[str] = "https://oauth2.googleapis.com/token"
    ):
        """
        Get or create Google Sheets API service.

        Args:
           access_token: Optional OAuth access token
           refresh_token: Optional OAuth refresh token
           client_id: OAuth client ID (required for refresh)
           client_secret: OAuth client secret (required for refresh)
           token_uri: Token URI for refreshing

        Returns:
            Google Sheets API service instance

        Raises:
            RuntimeError: If credentials are not configured
        """
        # Note: If we have a refresh token (even if we have an existing service), 
        # we might want to re-initialize to ensure we can refresh.
        # But for simplicity, if we have a service initialized with user creds, reuse it?
        # Ideally, we should check if the current service's creds match the requested ones.
        # For now, let's just always rebuild if access_token is provided to be safe/stateless for this request scope.
        
        try:
            from google.oauth2.credentials import Credentials
            from google.oauth2 import service_account
            from googleapiclient.discovery import build

            if access_token:
                # If we have refresh logic components
                if refresh_token and client_id and client_secret:
                    creds = Credentials(
                        token=access_token,
                        refresh_token=refresh_token,
                        token_uri=token_uri,
                        client_id=client_id,
                        client_secret=client_secret
                    )
                else:
                    # Access token only (might expire)
                    creds = Credentials(token=access_token)
                
                service = build('sheets', 'v4', credentials=creds)
                logger.info("✅ Google Sheets service initialized with User Token")
                return service

            if not self.google_credentials_path or not os.path.exists(self.google_credentials_path):
                raise RuntimeError(
                    "Google Sheets credentials not found. "
                    "Set GOOGLE_CREDENTIALS_PATH environment variable or provide credentials_path"
                )

            credentials = service_account.Credentials.from_service_account_file(
                self.google_credentials_path,
                scopes=['https://www.googleapis.com/auth/spreadsheets',
                       'https://www.googleapis.com/auth/drive.file']
            )

            self._sheets_service = build('sheets', 'v4', credentials=credentials)
            logger.info("✅ Google Sheets service initialized with Service Account")
            return self._sheets_service

        except ImportError:
            raise RuntimeError(
                "Google Sheets dependencies not installed. "
                "Run: pip install google-auth google-api-python-client"
            )
        except Exception as e:
            logger.error(f"Failed to initialize Google Sheets service: {e}")
            raise

    def export_to_google_sheets(
        self,
        data: List[Dict[str, Any]],
        spreadsheet_id: Optional[str] = None,
        sheet_name: str = "Facebook Ads Data",
        title: str = "Facebook Ads Analytics Export",
        access_token: Optional[str] = None,
        refresh_token: Optional[str] = None,
        client_id: Optional[str] = None,
        client_secret: Optional[str] = None
    ) -> str:
        """
        Export data to Google Sheets.

        Args:
            data: List of dicts to export (each dict is a row)
            spreadsheet_id: Optional existing spreadsheet ID
            sheet_name: Name of the sheet tab
            title: Spreadsheet title (for new spreadsheets)
            access_token: Optional OAuth access token
            refresh_token: Optional OAuth refresh token
            client_id: OAuth Client ID
            client_secret: OAuth Client Secret

        Returns:
            Spreadsheet URL
        """
        if not data:
            raise ValueError("No data to export")

        service = self._get_sheets_service(
            access_token=access_token,
            refresh_token=refresh_token,
            client_id=client_id,
            client_secret=client_secret
        )

        try:
            # Create new spreadsheet or use existing
            if not spreadsheet_id:
                spreadsheet_id = self._create_spreadsheet(service, title, sheet_name)

            # Prepare data for sheets
            headers = list(data[0].keys())
            rows = [[self._format_cell_value(row.get(header, "")) for header in headers]
                    for row in data]

            # Write headers
            self._write_to_sheet(service, spreadsheet_id, sheet_name, [headers], "A1")

            # Write data
            self._write_to_sheet(service, spreadsheet_id, sheet_name, rows, "A2")

            # Apply formatting
            self._format_sheet(service, spreadsheet_id, sheet_name, headers, len(rows) + 1)

            spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"
            logger.info(f"✅ Data exported to Google Sheets: {spreadsheet_url}")
            return spreadsheet_url

        except Exception as e:
            logger.error(f"Failed to export to Google Sheets: {e}")
            raise RuntimeError(f"Google Sheets export failed: {str(e)}")

    def _create_spreadsheet(self, service, title: str, sheet_title: str) -> str:
        """Create a new Google Spreadsheet"""
        spreadsheet = {
            'properties': {
                'title': title
            },
            'sheets': [
                {
                    'properties': {
                        'title': sheet_title
                    }
                }
            ]
        }
        spreadsheet = service.spreadsheets().create(
            body=spreadsheet,
            fields='spreadsheetId'
        ).execute()

        spreadsheet_id = spreadsheet.get('spreadsheetId')
        logger.info(f"Created new spreadsheet: {spreadsheet_id}")
        return spreadsheet_id

    def _write_to_sheet(
        self,
        service,
        spreadsheet_id: str,
        sheet_name: str,
        data: List[List[Any]],
        range_start: str
    ):
        """Write data to a sheet range"""
        # Quote sheet name if it contains spaces
        if " " in sheet_name:
            range_name = f"'{sheet_name}'!{range_start}"
        else:
            range_name = f"{sheet_name}!{range_start}"

        body = {
            'values': data
        }

        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name,
            valueInputOption='RAW',
            body=body
        ).execute()

    def _format_sheet(
        self,
        service,
        spreadsheet_id: str,
        sheet_name: str,
        headers: List[str],
        num_rows: int
    ):
        """
        Apply formatting to the sheet.

        - Bold headers
        - Freeze header row
        - Auto-resize columns
        - Format currency/percentage columns
        """
        # Get sheet ID
        sheet_metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = sheet_metadata.get('sheets', [])
        sheet_id = None

        for sheet in sheets:
            if sheet.get('properties', {}).get('title') == sheet_name:
                sheet_id = sheet.get('properties', {}).get('sheetId')
                break

        if sheet_id is None:
            logger.warning(f"Sheet '{sheet_name}' not found for formatting")
            return

        requests = []

        # Bold headers
        requests.append({
            'repeatCell': {
                'range': {
                    'sheetId': sheet_id,
                    'startRowIndex': 0,
                    'endRowIndex': 1
                },
                'cell': {
                    'userEnteredFormat': {
                        'textFormat': {
                            'bold': True
                        },
                        'backgroundColor': {
                            'red': 0.9,
                            'green': 0.9,
                            'blue': 0.9
                        }
                    }
                },
                'fields': 'userEnteredFormat(textFormat,backgroundColor)'
            }
        })

        # Freeze header row
        requests.append({
            'updateSheetProperties': {
                'properties': {
                    'sheetId': sheet_id,
                    'gridProperties': {
                        'frozenRowCount': 1
                    }
                },
                'fields': 'gridProperties.frozenRowCount'
            }
        })

        # Auto-resize columns
        requests.append({
            'autoResizeDimensions': {
                'dimensions': {
                    'sheetId': sheet_id,
                    'dimension': 'COLUMNS',
                    'startIndex': 0,
                    'endIndex': len(headers)
                }
            }
        })

        # Apply number formatting for common metric columns
        for col_idx, header in enumerate(headers):
            header_lower = header.lower()

            # Currency format
            if any(term in header_lower for term in ['spend', 'cost', 'value', 'cpc', 'cpa', 'cpm']):
                requests.append({
                    'repeatCell': {
                        'range': {
                            'sheetId': sheet_id,
                            'startRowIndex': 1,
                            'endRowIndex': num_rows,
                            'startColumnIndex': col_idx,
                            'endColumnIndex': col_idx + 1
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'numberFormat': {
                                    'type': 'CURRENCY',
                                    'pattern': '"$"#,##0.00'
                                }
                            }
                        },
                        'fields': 'userEnteredFormat.numberFormat'
                    }
                })

            # Percentage format
            elif any(term in header_lower for term in ['ctr', 'rate', 'roas', 'change']):
                requests.append({
                    'repeatCell': {
                        'range': {
                            'sheetId': sheet_id,
                            'startRowIndex': 1,
                            'endRowIndex': num_rows,
                            'startColumnIndex': col_idx,
                            'endColumnIndex': col_idx + 1
                        },
                        'cell': {
                            'userEnteredFormat': {
                                'numberFormat': {
                                    'type': 'NUMBER',
                                    'pattern': '#,##0.00'
                                }
                            }
                        },
                        'fields': 'userEnteredFormat.numberFormat'
                    }
                })

        # Execute all formatting requests
        if requests:
            body = {
                'requests': requests
            }
            service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            ).execute()

            logger.info(f"✅ Applied formatting to sheet '{sheet_name}'")

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str = "facebook_ads_export.xlsx",
        sheet_name: str = "Data"
    ) -> BytesIO:
        """
        Export data to Excel file (in-memory).

        Args:
            data: List of dicts to export
            filename: Filename for the Excel file
            sheet_name: Name of the Excel sheet

        Returns:
            BytesIO object containing the Excel file

        Example:
            ```python
            excel_file = export_service.export_to_excel(data)
            with open("export.xlsx", "wb") as f:
                f.write(excel_file.getvalue())
            ```
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

        except ImportError:
            raise RuntimeError(
                "Excel export dependencies not installed. "
                "Run: pip install openpyxl"
            )

        if not data:
            raise ValueError("No data to export")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = self._format_cell_value(row_data.get(header, ""))
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(header))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                try:
                    cell_length = len(str(row[0].value))
                    if cell_length > max_length:
                        max_length = cell_length
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        logger.info(f"✅ Data exported to Excel: {len(data)} rows")
        return excel_file

    def _format_cell_value(self, value: Any) -> Any:
        """Format a cell value for export"""
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, (dict, list)):
            import json
            return json.dumps(value)
        return str(value)
